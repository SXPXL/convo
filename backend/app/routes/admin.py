from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query
from sqlalchemy.orm import Session
import io
import csv
import re
from app.database import get_db
from app.auth.security import RoleChecker
from app.models.staff import Staff
from app.models.user import User
from app.models.entry import Entry
from app.schemas.staff import StaffCreate, StaffResponse
from app.schemas.user import UserCreate, UserResponse
from app.crud.staff import create_staff, get_staff_by_username
from app.crud.user import create_user, get_user_by_register_number, get_user_by_admission_number
from app.utils.qr import generate_qr_code_base64

router = APIRouter(prefix="/api/admin", tags=["admin"])

# Enforce admin only access
admin_auth = RoleChecker(allowed_roles=["admin"])

@router.post("/staff", response_model=StaffResponse, status_code=status.HTTP_201_CREATED)
def register_staff_member(
    staff_in: StaffCreate,
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    existing_staff = get_staff_by_username(db, staff_in.username)
    if existing_staff:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username already registered."
        )
    return create_staff(db, staff_in)

@router.get("/staff", response_model=list[StaffResponse])
def get_all_staff(
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    return db.query(Staff).all()

@router.delete("/staff/{staff_id}", status_code=status.HTTP_200_OK)
def delete_staff_member(
    staff_id: int,
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    if staff_id == current_staff.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="You cannot delete your own admin account."
        )
    
    staff = db.query(Staff).filter(Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Staff member not found."
        )
    
    db.delete(staff)
    db.commit()
    return {"detail": "Staff member deleted successfully."}

@router.post("/user", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def register_user(
    user_in: UserCreate,
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    existing_user = get_user_by_register_number(db, user_in.register_number)
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Register number already exists."
        )
        
    # If it's a guardian, verify linked student exists
    if user_in.type == "guardian" and user_in.linked_student_id:
        student = db.query(User).filter(User.id == user_in.linked_student_id).first()
        if not student:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Linked student with ID {user_in.linked_student_id} not found."
            )
            
    db_user = create_user(db, user_in)
    return db_user

@router.get("/user/{register_number}/qr")
def get_user_qr(
    register_number: str,
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    user = get_user_by_register_number(db, register_number)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"User with register number '{register_number}' not found."
        )
        
    qr_base64 = generate_qr_code_base64(user.register_number)
    return {
        "register_number": user.register_number,
        "name": user.name,
        "qr_code_url": qr_base64
    }

@router.post("/upload-participants")
def upload_participants(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    filename = file.filename.lower()
    content = file.file.read()
    
    rows = []
    try:
        if filename.endswith(".csv"):
            text_content = content.decode("utf-8-sig", errors="ignore")
            csv_reader = csv.DictReader(io.StringIO(text_content))
            headers = csv_reader.fieldnames
            normalized_headers = {h.strip().lower(): h for h in headers} if headers else {}
            
            for row in csv_reader:
                cleaned_row = {k.strip().lower(): v.strip() if v else "" for k, v in row.items() if k}
                rows.append(cleaned_row)
        elif filename.endswith(".xlsx"):
            import openpyxl
            wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
            sheet = wb.active
            header_row = [cell.value for cell in sheet[1]]
            normalized_headers = {}
            for idx, val in enumerate(header_row):
                if val is not None:
                    normalized_headers[str(val).strip().lower()] = idx
            
            for row_idx in range(2, sheet.max_row + 1):
                row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(header_row) + 1)]
                if not any(row_values):
                    continue
                cleaned_row = {}
                for k_lower, col_idx in normalized_headers.items():
                    if col_idx < len(row_values):
                        val = row_values[col_idx]
                        cleaned_row[k_lower] = str(val).strip() if val is not None else ""
                rows.append(cleaned_row)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported file format. Please upload a .csv or .xlsx file."
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse file: {str(e)}"
        )
        
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The uploaded file contains no data."
        )

    # Match column headers robustly
    keys = rows[0].keys()
    col_admission = None
    col_register = None
    col_student_name = None
    col_guardian_1 = None
    col_guardian_2 = None
    col_course = None
    
    for k in keys:
        if "2nd" in k or "second" in k or "guardian 2" in k or "parent 2" in k or ("participant" in k and ("2nd" in k or "2" in k)):
            col_guardian_2 = k
        elif "1st" in k or "first" in k or "guardian 1" in k or "parent 1" in k or ("participant" in k and ("1st" in k or "1" in k)):
            col_guardian_1 = k
        elif "admission" in k or "admn" in k or "adm no" in k:
            col_admission = k
        elif "register" in k or "reg" in k:
            col_register = k
        elif "course" in k or "dept" in k or "department" in k:
            col_course = k
        elif "student name" in k or "student_name" in k:
            col_student_name = k
            
    # Fallbacks for student name
    if not col_student_name:
        for k in keys:
            if "student" in k:
                col_student_name = k
                break
    if not col_student_name:
        for k in keys:
            if "name" in k and k != col_guardian_1 and k != col_guardian_2:
                col_student_name = k
                break
                
    if not col_register or not col_student_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Required columns not found. Detected columns: {list(keys)}. We need at least one column for Student Name and one for Register No."
        )

    imported_students = 0
    imported_guardians = 0
    errors = []
    
    for idx, r in enumerate(rows):
        try:
            reg_no = r.get(col_register, "").strip()
            name = r.get(col_student_name, "").strip()
            admn_no = r.get(col_admission, "").strip() if col_admission else ""
            course_val = r.get(col_course, "").strip() if col_course else ""
            g1_val = r.get(col_guardian_1, "").strip() if col_guardian_1 else ""
            g2_val = r.get(col_guardian_2, "").strip() if col_guardian_2 else ""
            
            # If only Guardian 2 is filled and Guardian 1 is blank, force Guardian 2 to occupy Guardian 1 (G1) slot
            if not g1_val and g2_val:
                g1_val = g2_val
                g2_val = ""
                
            if not reg_no or not name:
                continue
                
            # Clean department name: capitalize, remove text inside (), convert BSC/B.SC to BS
            dept_name = course_val
            dept_name = re.sub(r"\(.*?\)", "", dept_name)
            dept_name = dept_name.upper()
            dept_name = re.sub(r"\bBSC\b", "BS", dept_name)
            dept_name = re.sub(r"\bB\.SC\b", "BS", dept_name)
            dept_name = re.sub(r"\s+", " ", dept_name).strip()
                
            # Check if Student already exists in DB
            student = db.query(User).filter(User.register_number == reg_no).first()
            if not student:
                student = User(
                    register_number=reg_no,
                    admission_number=admn_no if admn_no else None,
                    name=name,
                    type="student",
                    department=dept_name if dept_name else None
                )
                db.add(student)
                db.commit()
                db.refresh(student)
                imported_students += 1
            else:
                if admn_no:
                    student.admission_number = admn_no
                if name:
                    student.name = name
                if dept_name:
                    student.department = dept_name
                db.commit()
                db.refresh(student)
                
            # Guardian 1
            if g1_val:
                g1_reg = f"{reg_no}-1"
                g1_admn = f"{admn_no}-1" if admn_no else None
                
                g1 = db.query(User).filter(User.register_number == g1_reg).first()
                if not g1:
                    g1 = User(
                        register_number=g1_reg,
                        admission_number=g1_admn,
                        name=g1_val,
                        type="guardian",
                        linked_student_id=student.id
                    )
                    db.add(g1)
                    imported_guardians += 1
                else:
                    g1.name = g1_val
                db.commit()
                
            # Guardian 2
            if g2_val:
                g2_reg = f"{reg_no}-2"
                g2_admn = f"{admn_no}-2" if admn_no else None
                
                g2 = db.query(User).filter(User.register_number == g2_reg).first()
                if not g2:
                    g2 = User(
                        register_number=g2_reg,
                        admission_number=g2_admn,
                        name=g2_val,
                        type="guardian",
                        linked_student_id=student.id
                    )
                    db.add(g2)
                    imported_guardians += 1
                else:
                    g2.name = g2_val
                db.commit()
                
        except Exception as row_error:
            db.rollback()
            errors.append(f"Row {idx+2}: {str(row_error)}")
            
    return {
        "status": "success" if not errors else "partial_success",
        "message": f"Successfully imported {imported_students} students and {imported_guardians} guardians.",
        "imported_students": imported_students,
        "imported_guardians": imported_guardians,
        "errors": errors
    }

@router.post("/reset-entries")
def reset_all_entries(
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    # Clear all records from Entries table
    db.query(Entry).delete()
    db.commit()
    return {"message": "All check-in entries have been successfully cleared."}

@router.get("/students")
def search_students(
    q: str = Query("", description="Search query by name or register number"),
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    query = db.query(User).filter(User.type == "student")
    if q:
        search_filter = f"%{q}%"
        query = query.filter(
            (User.name.ilike(search_filter)) | (User.register_number.ilike(search_filter))
        )
    students = query.limit(50).all()
    
    result = []
    for s in students:
        # Find guests linked to this student
        guests = db.query(User).filter(User.type == "guardian", User.linked_student_id == s.id).all()
        guest1 = next((g for g in guests if g.register_number.endswith("-1")), None)
        guest2 = next((g for g in guests if g.register_number.endswith("-2")), None)
        result.append({
            "id": s.id,
            "name": s.name,
            "register_number": s.register_number,
            "admission_number": s.admission_number,
            "department": s.department,
            "guest1_name": guest1.name if guest1 else "",
            "guest2_name": guest2.name if guest2 else ""
        })
    return result

@router.post("/students")
def create_student(
    payload: dict,
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    name = payload.get("name")
    register_number = payload.get("register_number")
    admission_number = payload.get("admission_number")
    department = payload.get("department")
    guest1_name = payload.get("guest1_name")
    guest2_name = payload.get("guest2_name")
    
    if not name or not register_number or not department:
        raise HTTPException(status_code=400, detail="Name, Register Number, and Department/Course are required.")
        
    # Check duplicate student register number
    existing = db.query(User).filter(User.register_number == register_number).first()
    if existing:
        raise HTTPException(status_code=400, detail="Student with this register number already exists.")
        
    student = User(
        name=name,
        register_number=register_number,
        admission_number=admission_number or None,
        department=department,
        type="student"
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    
    # Guest 1
    if guest1_name:
        g1_reg = f"{register_number}-1"
        g1_admn = f"{admission_number}-1" if admission_number else None
        # Clean any legacy left-over with this register number
        db.query(User).filter(User.register_number == g1_reg).delete()
        g1 = User(
            name=guest1_name,
            register_number=g1_reg,
            admission_number=g1_admn,
            type="guardian",
            linked_student_id=student.id
        )
        db.add(g1)
        
    # Guest 2
    if guest2_name:
        g2_reg = f"{register_number}-2"
        g2_admn = f"{admission_number}-2" if admission_number else None
        # Clean any legacy left-over with this register number
        db.query(User).filter(User.register_number == g2_reg).delete()
        g2 = User(
            name=guest2_name,
            register_number=g2_reg,
            admission_number=g2_admn,
            type="guardian",
            linked_student_id=student.id
        )
        db.add(g2)
        
    db.commit()
    return {"message": "Student created successfully.", "student_id": student.id}

@router.put("/students/{student_id}")
def update_student(
    student_id: int,
    payload: dict,
    db: Session = Depends(get_db),
    current_staff: Staff = Depends(admin_auth)
):
    student = db.query(User).filter(User.id == student_id, User.type == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found.")
        
    name = payload.get("name")
    register_number = payload.get("register_number")
    admission_number = payload.get("admission_number")
    department = payload.get("department")
    guest1_name = payload.get("guest1_name")
    guest2_name = payload.get("guest2_name")
    
    if not name or not register_number or not department:
        raise HTTPException(status_code=400, detail="Name, Register Number, and Department/Course are required.")
        
    # Check duplicate register number
    dup = db.query(User).filter(User.register_number == register_number, User.id != student_id).first()
    if dup:
        raise HTTPException(status_code=400, detail="Register number already in use by another user.")
        
    student.name = name
    student.register_number = register_number
    student.admission_number = admission_number or None
    student.department = department
    
    # Query current guests
    guests = db.query(User).filter(User.type == "guardian", User.linked_student_id == student_id).all()
    g1 = next((g for g in guests if g.register_number.endswith("-1")), None)
    g2 = next((g for g in guests if g.register_number.endswith("-2")), None)
    
    g1_reg = f"{register_number}-1"
    g1_admn = f"{admission_number}-1" if admission_number else None
    
    g2_reg = f"{register_number}-2"
    g2_admn = f"{admission_number}-2" if admission_number else None
    
    # Handle Guest 1
    if guest1_name:
        if g1:
            g1.name = guest1_name
            g1.register_number = g1_reg
            g1.admission_number = g1_admn
        else:
            # Clean any legacy left-over with this register number
            db.query(User).filter(User.register_number == g1_reg).delete()
            g1 = User(
                name=guest1_name,
                register_number=g1_reg,
                admission_number=g1_admn,
                type="guardian",
                linked_student_id=student_id
            )
            db.add(g1)
    else:
        if g1:
            db.delete(g1)
            
    # Handle Guest 2
    if guest2_name:
        if g2:
            g2.name = guest2_name
            g2.register_number = g2_reg
            g2.admission_number = g2_admn
        else:
            # Clean any legacy left-over with this register number
            db.query(User).filter(User.register_number == g2_reg).delete()
            g2 = User(
                name=guest2_name,
                register_number=g2_reg,
                admission_number=g2_admn,
                type="guardian",
                linked_student_id=student_id
            )
            db.add(g2)
    else:
        if g2:
            db.delete(g2)
            
    db.commit()
    return {"message": "Student updated successfully."}
